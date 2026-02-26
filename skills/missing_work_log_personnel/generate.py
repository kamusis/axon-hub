import argparse
import re
from collections import OrderedDict
from collections.abc import Sequence
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


EXCLUDE_NAMES = {"张乐奕", "杨廷琨", "黄宸宁", "李华", "李海清", "李轶楠", "章芋文"}
SPLIT_RE = re.compile(r"[、,，;；/|\n]+")


def _normalize_name(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = [p.strip() for p in SPLIT_RE.split(text) if p and p.strip()]
    return [p.replace(" ", "") for p in parts if p.replace(" ", "")]


def _header_map(values: Sequence[object]) -> dict[str, int]:
    m: dict[str, int] = {}
    for idx, v in enumerate(values):
        if v is None:
            continue
        key = str(v).strip()
        if key and key not in m:
            m[key] = idx
    return m


def _find_sheet_with_col(wb, col_name: str) -> tuple[str, int, int] | None:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx in range(1, min(ws.max_row, 30) + 1):
            row_values = [c.value for c in ws[row_idx]]
            hm = _header_map(row_values)
            if col_name in hm:
                return sheet_name, row_idx, hm[col_name] + 1
    return None


def _find_in_sheet(ws, col_name: str) -> tuple[int, int] | None:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        row_values = [c.value for c in ws[row_idx]]
        hm = _header_map(row_values)
        if col_name in hm:
            return row_idx, hm[col_name] + 1
    return None


def _read_column_values(ws, header_row: int, col_idx: int) -> list[str]:
    out: list[str] = []
    for r in range(header_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=r, column=col_idx).value
        out.extend(_normalize_name(cell_value))
    return out


def _parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    m = re.search(r"(\d{4})[/-\.](\d{1,2})[/-\.](\d{1,2})", text)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def _find_col_idx_by_candidates(hm: dict[str, int], candidates: list[str]) -> int | None:
    for name in candidates:
        if name in hm:
            return hm[name] + 1
    return None


def _read_report_people_and_span(
    ws,
    header_row: int,
    person_col_idx: int,
    start_col_idx: int | None,
    end_col_idx: int | None,
) -> tuple[list[str], date | None, date | None]:
    people_raw: list[str] = []
    min_start: date | None = None
    max_end: date | None = None
    for r in range(header_row + 1, ws.max_row + 1):
        people_raw.extend(_normalize_name(ws.cell(row=r, column=person_col_idx).value))
        if start_col_idx is not None:
            d = _parse_date(ws.cell(row=r, column=start_col_idx).value)
            if d is not None and (min_start is None or d < min_start):
                min_start = d
        if end_col_idx is not None:
            d = _parse_date(ws.cell(row=r, column=end_col_idx).value)
            if d is not None and (max_end is None or d > max_end):
                max_end = d
    people = _unique_preserve_order([p for p in people_raw if p])
    return people, min_start, max_end


def _read_active_rows(ws, header_row: int) -> tuple[list[str], list[list[object]]]:
    headers = [c.value for c in ws[header_row]]
    data: list[list[object]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        data.append(row)
    return [str(h).strip() if h is not None else "" for h in headers], data


def _unique_preserve_order(items: list[str]) -> list[str]:
    d: OrderedDict[str, None] = OrderedDict()
    for it in items:
        if it and it not in d:
            d[it] = None
    return list(d.keys())


def _write_output(
    output_path: Path,
    report_people: list[str],
    missing_rows_headers: list[str],
    missing_rows: list[list[object]],
):
    wb = Workbook()
    ws_missing = wb.active
    if ws_missing is None:
        ws_missing = wb.create_sheet("未报工人员名单")
    else:
        ws_missing.title = "未报工人员名单"
    ws_missing.append(missing_rows_headers)
    for row in missing_rows:
        ws_missing.append(row)

    ws_report = wb.create_sheet("已报工人员名单")
    ws_report.append(["姓名"])
    for name in report_people:
        ws_report.append([name])

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--active-file", default=None)
    parser.add_argument("--report-file", default=None)
    parser.add_argument("--input", default=None)
    parser.add_argument("--output", default=None)
    parser.add_argument("--active-sheet", default=None)
    parser.add_argument("--active-col", default="姓名")
    parser.add_argument("--report-sheet", default=None)
    parser.add_argument("--report-col", default="处理人")
    args = parser.parse_args()

    active_path = Path(args.active_file or args.input or "active_employee_list.xlsx")
    if not active_path.exists():
        print(f"Error: Active employee list not found at {active_path}")
        raise FileNotFoundError(str(active_path))

    report_path = Path(args.report_file) if args.report_file else None
    if report_path is None:
        default_report = Path("work_logs.xlsx")
        report_path = default_report if default_report.exists() else active_path
    if not report_path.exists():
        print(f"Error: Work log file not found at {report_path}")
        raise FileNotFoundError(str(report_path))

    print(f"Loading files...\n  Active list: {active_path}\n  Work logs: {report_path}")

    active_wb = load_workbook(active_path, data_only=True)
    report_wb = active_wb if report_path.resolve() == active_path.resolve() else load_workbook(report_path, data_only=True)

    if args.report_sheet:
        if args.report_sheet not in report_wb.sheetnames:
            raise ValueError(f"找不到工作表 {args.report_sheet}")
        report_ws = report_wb[args.report_sheet]
        loc = _find_in_sheet(report_ws, args.report_col)
        if loc is None:
            raise ValueError(f"在 {args.report_sheet} 中找不到列 {args.report_col}")
        report_header_row, report_col_idx = loc
    else:
        report_loc = _find_sheet_with_col(report_wb, args.report_col)
        if report_loc is None:
            raise ValueError(f"无法在任意工作表中找到列 {args.report_col}")
        report_sheet, report_header_row, report_col_idx = report_loc
        report_ws = report_wb[report_sheet]

    report_header_values = [c.value for c in report_ws[report_header_row]]
    report_hm = _header_map(report_header_values)
    start_idx = _find_col_idx_by_candidates(report_hm, ["开始时间", "开始日期", "开始"])
    end_idx = _find_col_idx_by_candidates(report_hm, ["结束时间", "结束时间时间", "结束", "结束日期"])
    report_people, span_start, span_end = _read_report_people_and_span(
        report_ws, report_header_row, report_col_idx, start_idx, end_idx
    )
    report_people_set = set(report_people)

    active_loc = None
    if args.active_sheet:
        if args.active_sheet not in active_wb.sheetnames:
            raise ValueError(f"找不到工作表 {args.active_sheet}")
        active_ws = active_wb[args.active_sheet]
        loc = _find_in_sheet(active_ws, args.active_col)
        if loc is not None:
            row_idx, col_idx = loc
            active_loc = args.active_sheet, row_idx, col_idx
        if active_loc is None:
            raise ValueError(f"在 {args.active_sheet} 中找不到列 {args.active_col}")
    else:
        active_loc = _find_sheet_with_col(active_wb, args.active_col)
    if active_loc is None:
        raise ValueError(f"无法在任意工作表中找到列 {args.active_col}")
    active_sheet, active_header_row, active_col_idx = active_loc
    active_ws = active_wb[active_sheet]

    headers, active_rows = _read_active_rows(active_ws, active_header_row)
    missing_rows: list[list[object]] = []

    desired_headers = ["姓名", "资源区域", "资源二级部门", "资源三级部门", "直接主管"]
    active_hm = _header_map(headers)
    desired_indices: list[int | None] = []
    for h in desired_headers:
        if h == "姓名":
            desired_indices.append(active_col_idx - 1)
        else:
            desired_indices.append(active_hm.get(h))

    for row in active_rows:
        if active_col_idx - 1 >= len(row):
            continue
        names = _normalize_name(row[active_col_idx - 1])
        if not names:
            continue
        name = names[0]

        if name in EXCLUDE_NAMES:
            continue
        if name not in report_people_set:
            projected: list[object] = []
            for idx in desired_indices:
                projected.append(row[idx] if idx is not None and idx < len(row) else None)
            missing_rows.append(projected)

    if span_start is None or span_end is None:
        raise ValueError("无法从报工明细中识别开始时间/结束时间并统计时间跨度")
    span_text = f"{span_start.strftime('%y-%m-%d')}-{span_end.strftime('%y-%m-%d')}"
    output_name = f"missing_work_log_report_{span_text}.xlsx"
    if args.output:
        out_arg = Path(args.output)
        out_dir = out_arg if out_arg.suffix.lower() != ".xlsx" else out_arg.parent
    else:
        out_dir = Path(".")
    output_path = (out_dir / output_name).resolve()
    print(f"Analysis complete:")
    print(f"  Total active employees checked: {len(active_rows)}")
    print(f"  Total unique reporters found: {len(report_people)}")
    print(f"  Missing reporters identified: {len(missing_rows)} (excluding {len(EXCLUDE_NAMES)} designated names)")
    print(f"Saving report to: {output_path}")
    _write_output(
        output_path=output_path,
        report_people=report_people,
        missing_rows_headers=desired_headers,
        missing_rows=missing_rows,
    )


if __name__ == "__main__":
    main()
