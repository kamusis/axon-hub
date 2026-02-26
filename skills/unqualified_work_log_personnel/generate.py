import argparse
import re
from collections.abc import Sequence
from datetime import date, datetime
from pathlib import Path
import json

from openpyxl import Workbook, load_workbook


SPLIT_RE = re.compile(r"[、,，;；/|\n]+")
DEFAULT_EXCLUDE_NAMES = {"张乐奕", "杨廷琨", "黄宸宁", "李华", "李海清", "李轶楠", "章芋文"}


def _header_map(values: Sequence[object]) -> dict[str, int]:
    m: dict[str, int] = {}
    for idx, v in enumerate(values):
        if v is None:
            continue
        key = str(v).strip()
        if key and key not in m:
            m[key] = idx
    return m


def _parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    m = re.search(r"(\d{4})[/-\.](\d{1,2})[/-\.](\d{1,2})", text)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _compact_text(text: str) -> str:
    return re.sub(r"\s+", "", text)


def _is_simple_description(text: str, rules: dict) -> bool:
    raw = _normalize_text(text)
    t = _compact_text(raw).lower()
    if not t:
        return True
    if len(t) <= 4:
        return True
    if "\n" in raw or re.search(r"\b\d+\s*[、.]", raw):
        return False
    if len(t) >= 25 and (raw.count("；") + raw.count(";") + raw.count("。") + raw.count(".") >= 1):
        return False
    
    simple_phrases = rules.get("simple_phrases", [])
    for p in simple_phrases:
        pp = _compact_text(p).lower()
        if t == pp or (t.startswith(pp) and len(t) <= 12) or (t.endswith(pp) and len(t) <= 12):
            return True
    
    # Check if description is just a technology name
    tech_names = rules.get("tech_names", [])
    if t in tech_names:
        return True

    generic_verbs = rules.get("generic_verbs", [])
    hits = sum(1 for v in generic_verbs if v in t)
    if hits >= 1 and len(t) <= 8:
        return True
    return False


def _load_rules(rules_path: Path) -> dict:
    if rules_path.exists():
        try:
            with open(rules_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: Failed to load rules from {rules_path}: {e}")
    return {
        "simple_phrases": [],
        "tech_names": [],
        "generic_verbs": []
    }


def _split_names(value: str) -> set[str]:
    raw = _normalize_text(value)
    if not raw:
        return set()
    parts = [p.strip() for p in SPLIT_RE.split(raw) if p and p.strip()]
    return {p.replace(" ", "") for p in parts if p.replace(" ", "")}


def _find_sheet_with_col(wb, col_name: str) -> tuple[str, int, int] | None:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx in range(1, min(ws.max_row, 30) + 1):
            hm = _header_map([c.value for c in ws[row_idx]])
            if col_name in hm:
                return sheet_name, row_idx, hm[col_name] + 1
    return None


def _find_col_idx_by_candidates(hm: dict[str, int], candidates: list[str]) -> int | None:
    for name in candidates:
        if name in hm:
            return hm[name] + 1
    return None


def _build_manager_map(active_file: Path) -> dict[str, str]:
    wb = load_workbook(active_file, data_only=True)
    loc = _find_sheet_with_col(wb, "姓名")
    if loc is None:
        raise ValueError("无法在在职人员名单中找到列 姓名")
    sheet, header_row, name_col = loc
    ws = wb[sheet]
    hm = _header_map([c.value for c in ws[header_row]])
    manager_col = hm.get("直接主管")
    if manager_col is None:
        raise ValueError("无法在在职人员名单中找到列 直接主管")
    manager_col += 1

    m: dict[str, str] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = _normalize_text(ws.cell(row=r, column=name_col).value)
        if not name:
            continue
        manager = _normalize_text(ws.cell(row=r, column=manager_col).value)
        if name not in m:
            m[name] = manager
    return m


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--report-file", default="work_logs.xlsx")
    parser.add_argument("--active-file", default="active_employee_list.xlsx")
    parser.add_argument("--output-dir", default=".")
    parser.add_argument("--sheet", default="报工内容不合格")
    parser.add_argument("--exclude", default="、".join(sorted(DEFAULT_EXCLUDE_NAMES)))
    parser.add_argument("--rules-file", default=None)
    args = parser.parse_args()

    report_path = Path(args.report_file)
    active_path = Path(args.active_file)
    if not report_path.exists():
        print(f"Error: Work log file not found at {report_path}")
        raise FileNotFoundError(str(report_path))
    if not active_path.exists():
        print(f"Error: Active employee list not found at {active_path}")
        raise FileNotFoundError(str(active_path))

    print(f"Loading files...\n  Work logs: {report_path}\n  Active list: {active_path}")
    manager_map = _build_manager_map(active_path)
    exclude_names = _split_names(args.exclude) if args.exclude is not None else set(DEFAULT_EXCLUDE_NAMES)
    
    rules_file = Path(args.rules_file) if args.rules_file else Path(__file__).parent / "rules.json"
    rules = _load_rules(rules_file)

    wb = load_workbook(report_path, data_only=True)
    loc = _find_sheet_with_col(wb, "备注")
    if loc is None:
        raise ValueError("无法在工时统计中找到列 备注")
    sheet_name, header_row, remark_col = loc
    ws = wb[sheet_name]
    headers = [c.value for c in ws[header_row]]
    hm = _header_map(headers)

    person_col = hm.get("处理人")
    if person_col is None:
        raise ValueError("无法在工时统计中找到列 处理人")
    person_col += 1

    start_col = _find_col_idx_by_candidates(hm, ["开始时间"])
    end_col = _find_col_idx_by_candidates(hm, ["结束时间", "结束时间时间"])
    if start_col is None or end_col is None:
        raise ValueError("无法在工时统计中找到开始/结束时间列")

    min_start: date | None = None
    max_end: date | None = None
    bad_rows: list[list[object]] = []

    for r in range(header_row + 1, ws.max_row + 1):
        sd = _parse_date(ws.cell(row=r, column=start_col).value)
        ed = _parse_date(ws.cell(row=r, column=end_col).value)
        if sd is not None and (min_start is None or sd < min_start):
            min_start = sd
        if ed is not None and (max_end is None or ed > max_end):
            max_end = ed

        person = _normalize_text(ws.cell(row=r, column=person_col).value)
        if person in exclude_names:
            continue

        remark = _normalize_text(ws.cell(row=r, column=remark_col).value)
        if not remark:
            continue
        if not _is_simple_description(remark, rules):
            continue

        manager = manager_map.get(person, "")
        row_values: list[object] = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        bad_rows.append([manager] + row_values)

    if min_start is None or max_end is None:
        raise ValueError("未能统计到报工时间跨度（开始/结束时间为空或无法解析）")

    bad_rows.sort(key=lambda x: (_normalize_text(x[1 + (person_col - 1)]), _parse_date(x[1 + (start_col - 1)]) or date.min))

    out_wb = Workbook()
    out_ws = out_wb.active
    if out_ws is None:
        out_ws = out_wb.create_sheet(args.sheet)
    else:
        out_ws.title = args.sheet

    out_headers = ["直接主管"] + [str(h).strip() if h is not None else "" for h in headers]
    out_ws.append(out_headers)
    for row in bad_rows:
        out_ws.append(row)

    span_text = f"{min_start.strftime('%y-%m-%d')}-{max_end.strftime('%y-%m-%d')}"
    output_name = f"unqualified_work_log_report_{span_text}.xlsx"
    output_path = (Path(args.output_dir) / output_name).resolve()
    print(f"Audit complete:")
    print(f"  Processed records: {ws.max_row - header_row}")
    print(f"  Identified unqualified records: {len(bad_rows)}")
    print(f"  Excluded {len(exclude_names)} personnel from audit.")
    print(f"Saving report to: {output_path}")
    out_wb.save(output_path)


if __name__ == "__main__":
    main()
